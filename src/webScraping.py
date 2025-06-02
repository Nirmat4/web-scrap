#╔══════════════════════════╗ 
#║ Importacion de librerias ║ 
#╚══════════════════════════╝ 
from selenium import webdriver
from selenium.webdriver.chrome.service import Service 
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
from selenium.common.exceptions import StaleElementReferenceException, TimeoutException, NoSuchElementException, ElementNotInteractableException, ElementClickInterceptedException
import random
import re
import time
from rich import print
import requests
import uuid
import os
import zipfile
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import undetected_chromedriver as uc

#╔══════════════════════════════════╗ 
#║ Configuración de Bright Data     ║ 
#╚══════════════════════════════════╝ 
BRIGHTDATA_HOST = "brd.superproxy.io"
BRIGHTDATA_PORT = 22225  # Usar puerto HTTP para evitar problemas con SSL
BRIGHTDATA_USER = "brd-customer-hl_dbcb7806-zone-test_scrap"
BRIGHTDATA_PASS = "gk1x6bsjm4x9"

#╔══════════════════════════════════╗ 
#║ Funciones mejoradas              ║ 
#╚══════════════════════════════════╝ 
def create_proxy_extension_v3(host, port, user, pwd, plugin_path=None):
    import zipfile, json, random, string

    if plugin_path is None:
        rnd = "".join(random.choices(string.ascii_letters+string.digits, k=8))
        plugin_path = f"./proxys/proxy_auth_plugin_{rnd}.zip"

    manifest = {
      "name": "Proxy Auth Extension",
      "version": "1.0.0",
      "manifest_version": 3,
      "minimum_chrome_version": "108.0.0",
      "permissions": [
        "proxy",
        "storage",
        "unlimitedStorage",
        "webRequest",
        "webRequestAuthProvider"
      ],
      "host_permissions": ["<all_urls>"],
      "background": {"service_worker": "background.js"}
    }

    background_js = f"""
chrome.proxy.settings.set({{
  value: {{
    mode: "fixed_servers",
    rules: {{
      singleProxy: {{
        scheme: "http",
        host: "{host}",
        port: parseInt({port})
      }},
      bypassList: ["<local>"]
    }}
  }},
  scope: "regular"
}}, function(){{}});

chrome.webRequest.onAuthRequired.addListener(
  function(details) {{
    return {{authCredentials: {{username: "{user}", password: "{pwd}"}}}};
  }},
  {{urls: ["<all_urls>"]}},
  ["blocking"]
);
"""

    with zipfile.ZipFile(plugin_path, 'w') as zp:
        zp.writestr("manifest.json", json.dumps(manifest, indent=2))
        zp.writestr("background.js", background_js)

    return plugin_path

#╔══════════════════════════════════╗ 
#║ Función init_web modificada      ║ 
#╚══════════════════════════════════╝ 
def init_web():
    options = uc.ChromeOptions()
    
    proxy_ext = create_proxy_extension_v3(
        BRIGHTDATA_HOST, BRIGHTDATA_PORT,
        BRIGHTDATA_USER, BRIGHTDATA_PASS
    )
    options.add_extension(proxy_ext)

    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--window-size=1366,768")
    options.add_argument("--disable-blink-features=AutomationControlled")

    user_agent = (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/114.0.5735.199 Safari/537.36"
    )
    options.add_argument(f"user-agent={user_agent}")

    driver = uc.Chrome(options=options, use_subprocess=True)

    driver.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument", {
        "source": """
            Object.defineProperty(navigator, 'webdriver', {get: () => undefined});
            window.chrome = {runtime: {}};
            Object.defineProperty(navigator, 'permissions', {
                get: () => ({
                    query: (parameters) => (
                        parameters.name === 'notifications' ?
                        Promise.resolve({ state: Notification.permission }) :
                        Promise.resolve({ state: 'denied' })
                    )
                }),
            });
        """
    })

    try:
        driver.get("https://www.autocompara.com")
        WebDriverWait(driver, 20).until_not(
            lambda d: any(kw in d.title.lower() for kw in ["cloudflare", "captcha", "error"])
        )# tras cargar la página y antes de rellenar...
        csrf = None
        for cookie in driver.get_cookies():
            if cookie['name'] in ('XSRF-TOKEN','sec_cpt'):
                csrf = cookie['value']
                break
        print("CSRF token:", csrf)
        if csrf:
          driver.execute_cdp_cmd("Network.enable", {})
          driver.execute_cdp_cmd(
              "Network.setExtraHTTPHeaders",
              {"headers": {
                  "X-XSRF-TOKEN": csrf,
                  "Referer": "https://www.autocompara.com/",
                  # (o "If-Match": etag) si ese fuera el caso
              }}
          )

    except Exception as e:
        print(f"[red]Error durante navegación: {e}[/]")
        driver.quit()
        return None

    try:
        WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.ID, "sec-overlay")))
        driver.execute_script("document.getElementById('sec-overlay').remove()")
        print("[green]Overlay de seguridad eliminado[/]")
    except Exception as e:
        print(f"[yellow]No se pudo eliminar overlay: {str(e)}[/]")

    return driver



#╔════════════════════════════════════════════════════════╗
#║ Funcion para insertar el año del auto                  ║
#║  ° Posicionamiento del driver en el input              ║
#║  ° Digitamos el año del modelo (caracter por caracter) ║
#╚════════════════════════════════════════════════════════╝
def insert_year(driver, año: int):
  try:
    input_year=WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.ID, "year")))

    # -- Simulamos una espera de tiempo natural --
    time.sleep(0.5)

    # -- Limpiamos los valores default --
    input_year.clear()

    # -- Introducimos el año caracter a caracter --
    for char in str(año):
      input_year.send_keys(char)
      # -- Pausa entre caracteres --
      time.sleep(0.1)

    print(f"[bold cyan]año {año} insertado correctamente[/]")
    return True
  except TimeoutException:
    print("[bold red]tiempo de espera agotado, no se encontro el elemento con id 'year'[/]")
    return False
  except NoSuchElementException:
    print("[bold red]el elemento con id 'year' no existe en la pagina[/]")
    return False
  except Exception as e:
    print(f"[bold red]error inesperado: {str(e)}[/]")
    return False

#╔════════════════════════════════════════════════════════╗
#║ Funcion para busqueda de modelo                        ║
#║  ° Posicionamiento del driver en el input              ║
#║  ° Digitamos el año del modelo (caracter por caracter) ║
#╚════════════════════════════════════════════════════════╝
def insert_model(driver, texto: str):
  try:
    input_container=WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "div.ng-input")))

    # -- Centrado de elemento para evitar sobre ajuste --
    driver.execute_script(
      """arguments[0].scrollIntoView({
        behavior: 'auto',
        block: 'center',
        inline: 'center'
      });""",
      input_container
    )
    time.sleep(0.5)

    # -- Hacemos click en el elemento --
    driver.execute_script("arguments[0].click();", input_container)

    # -- Localizamos el input --
    input_field=WebDriverWait(driver, 5).until(
      EC.element_to_be_clickable((By.CSS_SELECTOR, "div.ng-input input"))
    )

    # -- Escritura con validacion de estado y formato --
    value=texto.upper() 
    for char in texto:
      input_field.send_keys(char)
      time.sleep(random.uniform(0.1, 0.25))
      current_value=input_field.get_attribute("value").upper()
      if char not in current_value:
        raise ValueError(f"Caracter '{char}' no detectado. Valor actual: {current_value}")
      
    WebDriverWait(driver, 10).until(lambda d: texto in d.find_element(By.CSS_SELECTOR, "div.ng-input input").get_attribute("value").upper())

    print(f"[bold cyan]busqueda exitosa para: {value}[/]")
    return True
  
  except Exception as e:
    print(f"[bold red]fallo en busqueda: {str(e)}")
    try:
      driver.execute_script("""
        document.querySelector('ng-select#search').click();
        document.querySelector('ng-select#search input').value=arguments[0];
        document.querySelector('ng-select#search input').dispatchEvent(new Event('input'));
      """, texto)
      print(f"[bold yellow]usando fallback JS para: {texto}[/]")
      return True
    except:
      return False
    
#╔════════════════════════════════════════════════════════╗
#║ Funcion para la seleccion del modelo de auto           ║
#║  ° Posicionamiento del driver en el select             ║
#║  ° Compara el texto del modelo con la opcion           ║
#║  ° Selecciona la opcion si llega a coincidir           ║
#╚════════════════════════════════════════════════════════╝
def select_model(driver, texto: str):
  try:
    WebDriverWait(driver, 15).until(EC.visibility_of_element_located((By.CSS_SELECTOR, "ng-dropdown-panel[id^='a']")))

    # -- Localizamos el contenedor scrollable --
    options=driver.find_elements(By.CSS_SELECTOR, "div.ng-option-child:not(.ng-optgroup)")

    # -- Buscamos la opcion deseada con scroll --
    find=False
    for option in options:
      try:
        driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", option)
        # -- Optenemos el texto ignorando elementos hijos --
        texto_option=option.find_element(By.CSS_SELECTOR, "div.vehicle-data-home").get_attribute("textContent").strip()
        if texto_option.lower()==texto.lower():
          # -- Click para evitar problemas de overlay --
          driver.execute_script("arguments[0].click();", option)
          print(f"[bold cyan]opcion seleccionada: {texto_option}[/]")
          find=True
          break
      except Exception as e:
        print(f"[bold red]error durante iteracion de opciones: {str(e)}[/]")
        continue
    if not find:
      # -- Usamos una opcion default
      print("[bold yellow]Opcion no encontrada, usando alternativa...[/]")
      options[0].click()
    return True
  except Exception as e:
    print(f"[bold red]error general: {str(e)}[/]")
    return False
  
#╔════════════════════════════════════════════════════════╗
#║ Funcion para la seleccion del modelo de auto default   ║
#║  ° Posicionamiento del driver en el select             ║
#║  ° Selecciona la opcion                                ║
#╚════════════════════════════════════════════════════════╝
def seleccionar_opcion_alternativa(driver, texto: str):
  try:
    option=WebDriverWait(driver, 10).until(EC.element_to_be_clickable((
      By.XPATH,
      f"//div[contains(@class, 'ng-option') and not(contains(@class, 'ng-optgroup'))]//div[contains(@class, 'vehicle-data-home') and normalize-space()='{texto}']/ancestor::div[contains(@class, 'ng-option')]"
    )))

    # -- Scroll a elemento y click --
    driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center', inline: 'nearest'});", option)
    time.sleep(0.3)
    driver.execute_script("arguments[0].click();", option)
    return True
  except Exception as e:
    print(f"[bold red]error en metodo alternativo: {str(e)}[/]")
    return False
  
#╔════════════════════════════════════════════════════════╗
#║ Funcion de uso unico para click de boton "continuar "  ║
#║  ° Posicionamiento del driver en el botton             ║
#║  ° Click                                               ║
#╚════════════════════════════════════════════════════════╝
def click_continuar(driver):
  try:
    boton=WebDriverWait(driver, 15).until(EC.element_to_be_clickable((
      By.XPATH, 
      "//button[contains(@class, 'btn-gradient') and .//text()[contains(., 'Continuar')]]"
    )))

    # -- Scroll al bton para asegurar visibilidad --
    driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});", boton)
    driver.execute_script("arguments[0].click();", boton)

    print("[bold cyan]boton continuar clickeado exitosamente[/]")
    return True
  except Exception as e:
    print(f"[bold red]error al hacer clic en continuar: {str(e)}[/]")
    return False
  

#############################################################
# LAS SIGUIENTES FUNCIONES CORRESPONDEN AL LLENADO DE DATOS #
# DE LOS FORMULARIOS DE DATOS PERSONALES                    #
#############################################################

#╔════════════════════════════════════════════════════════╗
#║ Funcion de uso unico para seleccion de genero          ║
#║  ° Posicionamiento del driver en el botton             ║
#║  ° Click                                               ║
#╚════════════════════════════════════════════════════════╝
def select_gender(driver, genero: str):
  try:
    selectores={
      "Hombre": {
        "xpath": "//label[contains(@class, 'male') and normalize-space()='Hombre']",
        "css": "label.male"
      },
      "Mujer": {
        "xpath": "//label[contains(@class, 'female') and normalize-space()='Mujer']",
        "css": "label.female"
      }
    }
    if genero not in selectores:
      raise ValueError(f"[bold purple]opcion invalida: {genero}. Use 'Hombre' o 'Mujer'[/]")
    
    # -- Localizacion de elemento --
    elemento=WebDriverWait(driver, 10).until(EC.element_to_be_clickable((
      By.XPATH, 
      selectores[genero]["xpath"]
    )))

    # -- Scroll y Click para evitar problemas de overlay
    driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});", elemento)
    driver.execute_script("arguments[0].click();", elemento)

    # -- Verificar selecion --
    input_radio=elemento.find_element(By.TAG_NAME, "input")
    if input_radio.is_selected():
      print(f"[bold cyan]genero seleccionado correctamente {genero}[/]")
      return True
    else:
      print(f"[bold red]error, el radio button no se marco[/]")
      return False
  except Exception as e:
    print(f"[bold red]error seleccionando genero: {str(e)}[/]")
    # -- Buscamos el boton por css --
    try:
      elemento=driver.find_element(By.CSS_SELECTOR, selectores[genero]["css"])
      elemento.click()
      return True
    except:
      print("[bold red]error con el selector en css[/]")
      return False
    
#╔════════════════════════════════════════════════════════╗
#║ Funcion de ingreso de fecha de nacimiento              ║
#║  ° Posicionamiento del driver en el input              ║
#║  ° Digitamos la fecha de nacimiento por caracter       ║
#║  * NOTA: Importante reformatear la fecha               ║
#╚════════════════════════════════════════════════════════╝
def insert_date(driver, fecha: str):
  try:
    campo_fecha=WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.ID, "date")))

    # -- Limpieza de campo de valores basura --
    driver.execute_script("arguments[0].value='';", campo_fecha)
    campo_fecha.clear()
    time.sleep(0.5)

    for i, char in enumerate(fecha):
      campo_fecha.send_keys(char)
      current_value=campo_fecha.get_attribute("value")
      # -- Validamos el formato mientras se escribe
      if i in [2, 5]:
        if len(current_value) <= i or current_value[i] != "/":
          driver.save_screenshot(f"error_formato_{i}.png")
          raise ValueError(f"[bold red]Formato invalido en posicion {i+1}")
      time.sleep(0.1)

    WebDriverWait(driver, 5).until(lambda d: d.find_element(By.ID, "date").get_attribute("value").replace("/", "") == fecha.replace("/", ""))

    print(f"[bold cyan]fecha ingresada correctamente {fecha}[/]")
    return True
  except ElementNotInteractableException:
    # -- El plan B es usar JavaScript para forzar la entrada --
    driver.execute_script(
      """
      const input=document.getElementById('date');
      input.value=arguments[0];
      input.dispatchEvent(new Event('input', { bubbles: true }));
      input.dispatchEvent(new Event('change', { bubbles: true }));
      """,
      fecha 
    )
    print(f"[bold purple]fecha ingresada mediante JavaScript[/]")
    return True
  except Exception as e:
    print(f"[bold red]error ingresando fecha: {str(e)}")
    driver.save_screenshot("error_fecha.png")
    return False
  
#╔════════════════════════════════════════════════════════╗
#║ Funcion de llenado de datos personales                 ║
#║  ° Posicionamiento del driver en el form               ║
#║  ° Digitamos nombre, cp, correo y numero por caracter  ║
#╚════════════════════════════════════════════════════════╝
def insert_personal_data(driver, nombre: str, codigo_p: str, email: str, telefono: str):
  try:
    campos={
      "name": {
        "valor": nombre,
        "validacion": lambda x: len(x.split()) >= 1
      },
      "cp": {
        "valor": codigo_p,
        "validacion": lambda x: x.isdigit() and len(x) == 5
      },
      "email": {
        "valor": email,
        "validacion": lambda x: re.match(r"[^@]+@[^@]+\.[^@]+", x)
      },
      "phone": {
        "valor": telefono,
        "validacion": lambda x: x.isdigit() and len(x) in [10, 12]
      }
    }

    # -- Llenado de cada campo caracter a caracter --
    for campo_id, config in campos.items():
      # -- Validacion de formato --
      if not config["validacion"](config["valor"]):
        raise ValueError(f"[bold red]formato invalido para {campo_id}: {config['valor']}")
      # -- Limpieza de campo
      elemento=WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.ID, campo_id)))
      elemento.clear()

      # -- Simulacion de escritura humana --
      for char in config["valor"]:
        elemento.send_keys(char)
       me.sleep(0.05) ti

    print("[bold cyan]todos los campos llenados correctamente")
    return True
  except Exception as e:
    print(f"[bold red]error en llenado de datos: {str(e)}")
    driver.save_screenshot("error_datos_personales.png")
    return False
  
#╔════════════════════════════════════════════════════════╗
#║ Funcion de uso unico para click de cotizacion          ║
#║  ° Posicionamiento del driver en el botton             ║
#║  ° Click para la generacion de la cotizacion           ║
#╚════════════════════════════════════════════════════════╝
def click_cotizar(driver):
  try:
    boton=WebDriverWait(driver, 25).until(EC.element_to_be_clickable((
      By.XPATH, 
      "//button[contains(@class, 'btn-gradient') and contains(., 'Cotizar')]"
    )))

    # -- Simulacion de uso Humano --
    driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});", boton)
    time.sleep(random.uniform(0.5, 1.5))

    # -- Click con Java script para evitar interceptaciones --
    driver.execute_script("arguments[0].click();", boton)
    print("[bold cyan]click en cotizar realizado[/]")

    # -- Esperamos la carga prolongada (conbinacion de enfoques)
    try:
      WebDriverWait(driver, 40).until(lambda d: any([
        len(d.find_element(By.CSS_SELECTOR, "div.resultado-cotizacion"))>0,
        "resultado" in d.current_url,
        d.execute_script("return document.readyState")=="complete"
      ]))
    except:
      # -- Espera de respaldo --
      time.sleep(3)
      if "cotizacion" not in driver.current_url:
        raise TimeoutError("[bold red] la pagina no termino de cargar[/]")
      if "cotizacion-error" in driver.current_url:
        print(f"[bold red]error con el envio de la informacion (hiciste demasiadas solicitudes con tu ip)[/]")
        return False
    
    print("[bold cyan]pagina de resultados cargada exitosamente")
    return True
  except Exception as e:
    print(f"[bold red]error final: {str(e)}[/]")
    # -- Diagnostico de error --
    print("[bold purple]estado actual:[/]")
    print(f"[bold purple]- URL: {driver.current_url}[/]")
    print(f"[bold purple]- titulo: {driver.title}[/]")
    print(f"[bold purple]- ReadyState: {driver.execute_script('return document.readyState')}[/]")
    driver.save_screenshot("error_final.png")
    return False
  
#╔════════════════════════════════════════════════════════╗
#║ Funcion para apertura de modal y observar informacion  ║
#║  ° Posicionamiento del driver en el label              ║
#║  ° Click para la apertura del modal                    ║
#╚════════════════════════════════════════════════════════╝
def open_modal(driver):
  try:
    WebDriverWait(driver, 15).until(lambda d: d.execute_script("return document.readyState") == "complete")

    # -- Localizamos el boton para el despliegue del modal --
    boton_selector=("button.btn.btn-lnk.details[acgtmevent][category='compara_y_elige'][action='informacion_seguros']")

    # -- Scroll al boton (por seguridad) --
    boton=WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.CSS_SELECTOR, boton_selector)))
    driver.execute_script(
      "arguments[0].scrollIntoView({behavior: 'auto', block: 'center', inline: 'center'});"
      "window.scrollBy(0, -window.innerHeight * 0.1);", 
      boton
    )
    time.sleep(random.uniform(0.3, 0.7))

    # -- click con metodos de respaldo --
    try:
      ActionChains(driver)\
        .move_to_element_with_offset(boton, 5, 5)\
        .pause(random.uniform(0.1, 0.3))\
        .click()\
        .perform()
    except:
      # -- Fallback con JavaScript --
      driver.execute_script("arguments[0].click();", boton)
    
    # -- Espera del modal --
    WebDriverWait(driver, 15).until(EC.visibility_of_element_located((By.CSS_SELECTOR, "div.modal-content")))
    WebDriverWait(driver, 10).until(lambda d: "coberturas" in d.page_source.lower())

    print("[bold cyan]modal de coberturas abierto correctamente[/]")
    return True
  except TimeoutException:
    print("[bold red]error tiempo de espera execido para el modal[/]")
    driver.save_screenshot("error_modal_timeout.png")
    return False
  except Exception as e:
    print(f"[bold red]error al abrir el modal: {str(e)}[/]")
    driver.save_screenshot("error_modal_general.png")

    # -- intentar cerrar overlays inesperados --
    try:
      driver.execute_script("""
        const modals=document.querySelectorAll('div.modal-backdrop');
        modals.forEach(modal => modal.remove());
        document.body.classList.remove('modal-open');
      """)
    except:
      pass

    return False
  
#╔═════════════════════════════════════════════════════════════════╗
#║ Funcion de ejecucion principal para evitar llamadas prolongadas ║
#║  ° Recibe toda la informacion para la ejecucion principal       ║
#║  ° retorna el driver y deja el navegador listo para la          ║
#║    extraccion principal de la informacion                       ║
#╚═════════════════════════════════════════════════════════════════╝
def init_browser():
  driver=init_web()
  return driver

def insert_auto(driver, ano, modelo):
  insert_year(driver, ano)
  insert_model(driver, modelo)
  select_model(driver, modelo)
  click_continuar(driver)
  return True

def insert_data(driver, genero, nacimiento, nombre, cp, email, telefono):
  select_gender(driver,genero)
  insert_date(driver, nacimiento)
  insert_personal_data(driver, nombre, cp, email, telefono)
  click_cotizar(driver)
  open_modal(driver)
  return True

"""
Ejemplo de ejecucion
init_browser(2015, "AVEO LS A STD 1.6L 4CIL 4PTAS", "Hombre", "16072002", "Emilio", "52977", "foyagev912@ofular.com", "524385654784", "51.81.245.3:17981")
"""

#############################################################
# LAS SIGUIENTES FUNCIONES CORRESPONDEN A LA EXTRACCION DE  #
# LA INFORMACION DE LA CONTIZACION.                         #
#############################################################

#╔═══════════════════════════════════════════════════════════╗
#║ Funcion para interactuar con el componente ng-select para ║
#║ desplegar sus opciones.                                   ║
#║  ° Toma intentos para la interaccion y poscionamiento     ║
#║  ° Despliega las opciones de seguros                      ║
#╚═══════════════════════════════════════════════════════════╝
def drop_options(driver, intentos=3):
  try:
    selector_ng_select="ng-select[name='insuaranse'].custom"
    ng_select=WebDriverWait(driver, 15).until(EC.presence_of_element_located((By.CSS_SELECTOR, selector_ng_select)))

    # -- Localizamos el area clickable --
    trigger_selector="div.ng-input[role='combobox']"
    trigger=ng_select.find_element(By.CSS_SELECTOR, trigger_selector)

    # -- Scroll preciso y seguro (evitar rastreo) --
    driver.execute_script("""
      const element=arguments[0];
      const header=document.querySelector('header') || { offsetHeight: 0 };
      const yOffset=-header.offsetHeight * 0.15;
      
      element.scrollIntoView({
          behavior: 'auto',
          block: 'center',
          inline: 'nearest'
      });
      
      window.scrollBy(0, yOffset);
    """, trigger)

    # -- Secuencia de iteracion robusta
    for intento in range(intentos):
      try:
        # -- Click con Actions para maxima precision --
        ActionChains(driver)\
          .move_to_element_with_offset(trigger, 5, 5)\
          .pause(0.25)\
          .click()\
          .pause(0.5)\
          .perform()
        
        # -- Verificar apertura --
        if trigger.get_attribute("aria-expanded") == "true":
          print("[bold cyan]dropdown desplegado[/]")
          return True
        
        # -- Fellback verificar clase de estado abierto --
        if "ng-select-opened" in ng_select.get_attribute("class"):
          print("[bold cyan]dropdown abierto (verificacion por clase)[/]")
          return True
        
        print(f"[bold yellow]intento {intento+1} - drop no abierto[/]")
      except StaleElementReferenceException:
        print("[bold yellow]elemento obsoleto, reintentando...[/]")
        trigger=ng_select.find_element(By.CSS_SELECTOR, trigger_selector)
        continue
    
    # -- Verificacion final del panel --
    panel_selector="ng-dropdown-panel.custom"
    if EC.visibility_of_element_located((By.CSS_SELECTOR, panel_selector))(driver):
      print("[bold cyan]panel detectado post-intentos[/]")
      return True
    
    raise Exception("[bold red]no se pudo desplegar el dropdown")
  except Exception as e:
    print(f"[bold red]error critico: {str(e)}")
    driver.save_screenshot("error_ng_select.png")
    return False

#╔══════════════════════════════════════════════════════════╗
#║ Funcion para obtener las aseguradoras (id y nombre)      ║
#║  ° Recupera todos los ids disponibles junto al nombre    ║
#╚══════════════════════════════════════════════════════════╝
def get_options(driver):
  try:
    panel_selector="ng-dropdown-panel.custom"
    panel=WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.CSS_SELECTOR, panel_selector)))

    # -- Obtener todos los elementos --
    options=panel.find_elements(By.CSS_SELECTOR, "div.ng-option[id]")

    # -- Procesamiento de opciones --
    options_data=[]
    for option in options:
      try:
        option_id=option.get_attribute("id")
        name=option.find_element(By.CSS_SELECTOR, "span.ng-option-label").text.strip()

        # -- Validamos y agregamos --
        if option_id and "-" in option_id and name:
          options_data.append((option_id, name))

      except Exception as e:
        print(f"[bold yellow]error en la opcion: {str(e)}[/]")
        continue
    print(f"[bold cyan]{len(options_data)} opciones encontradas")

    return options_data
  except Exception as e:
    print(f"[bold red]error general: {str(e)}[/]")
    driver.save_screenshot("error_opciones_completas.png")
    return []
  
#╔═══════════════════════════════════════════════════════════╗
#║ Funcion para cambiar de aseguradora por medio de un id    ║
#║  ° Toma el id de entrada y se posciona sobre el elemento  ║
#║  ° Selecciona el elemento siempre y cuando el id coincida ║
#╚═══════════════════════════════════════════════════════════╝
def change_option(driver, option_id: str):
  try:
    partial_id=option_id.split("-")[-1]

    # -- Verificacion del dropdown --
    panel_selector="ng-dropdown-panel.custom"
    for _ in range(2):
      # -- Verificacion por intentos --
      try:
        WebDriverWait(driver, 5).until(EC.visibility_of_element_located((By.CSS_SELECTOR, panel_selector)))
        break
      except:
        print("[bold yellow]reabriendo dropdown...[/]")
        if not drop_options(driver):
          raise Exception("[bold red]no se pudo abrir el dropdown[/]")
    
    # -- Busqueda por xpath para navegadores que si lo soporten --
    xpath_selector=f"//div[@role='option' and substring(@id, string-length(@id) - {len(partial_id)} + 1)='{partial_id}']"
    option=WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, xpath_selector)))

    # -- Scroll al elemento --
    ActionChains(driver)\
      .move_to_element(option)\
      .perform()
    time.sleep(0.2)

    # -- Reintentos --
    for intento in range(3):
      try:
        driver.execute_script("arguments[0].scrollIntoViewIfNeeded()", option)
        option.click()

        # -- Verificacion alternativa --
        WebDriverWait(driver, 2).until(EC.invisibility_of_element_located((By.CSS_SELECTOR, panel_selector)))
        print(f"[bold cyan]seleccion confirmada {option_id}[/]")
        return True
      except StaleElementReferenceException:
        print(f"[bold yellow]reintento {intento+1} - actualizando referencia...[/]")
        option=WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.XPATH, xpath_selector)))
      except Exception as e:
        print(f"[bold yellow]error en el intento {intento+1}: {str(e)}[/]")
        # -- intento por medio de JavaScript --
        driver.execute_script("arguments[0].click()", option)
    raise Exception("[bold red]maximo de intentos alcanzado[/]")
  except Exception as e:
    print(f"[bold red]fallo en {option_id}: {str(e)}[/]")
    driver.save_screenshot(f"error_{option_id.replace('-', '_')}.png")
    return False

#╔═══════════════════════════════════════════════════════════╗
#║ Funcion para la optencion del precio anual                ║
#║  ° Posicionamiento sobre elementos <p> sin hijos          ║
#║  ° Excluimos elementos ValueB                             ║
#║  ° Recuperamos el precio del seguro                       ║
#╚═══════════════════════════════════════════════════════════╝
def get_price(driver) -> str:
  try:
    xpath_selector=(
      ".//p[contains(@class, 'value') "
      # -- Excluimos elementos con hijos --
      "and not(*[not(self::text())]) "
      "and contains(., '$')]"
    )

    # -- Esperamos que los elementos cumplan con los criterior --
    precios=WebDriverWait(driver, 15).until(
      lambda d: [
        p for p in d.find_elements(By.XPATH, xpath_selector)
        if p.is_displayed() 
        and not {'valueB'}.intersection(p.get_attribute("class").split())
        and re.search(r"\$\s*\d", p.text)
      ]
    )

    # -- Priorizacion --
    elemento_correcto=None
    for p in precios:
      classes=p.get_attribute("class").split()

      # -- Primera evaluacion --
      if "valueA" in classes:
        elemento_correcto=p
        break
      if "value" in classes and not {"valueA", "valueB"}.intersection(classes):
        elemento_correcto=p
        break

    # -- Fallback controlado --
    if not elemento_correcto:
      elemento_correcto=precios[0] if precios else None
      print("[bold yellow]usando fallback de precio[/]")

    # -- Validacion final de estrcutura --
    if not elemento_correcto or elemento_correcto.find_elements(By.XPATH, "./*"):
      raise ValueError("[bold red]estrcutra de precio invalida[/]")
    
    # -- Extraccion por medio de regex --
    precio_texto=elemento_correcto.text
    match=re.search(
      r"""
      \$
      \s*
      (
          \d{1,3}
          (?:,\d{3})*
          (?:\.\d{2})?
      )
      """, 
      precio_texto, 
      re.VERBOSE
    )
    if not match:
      raise ValueError(f"[bold red]formato numerico invalido. {precio_texto}[/]")
    
    # -- Convertimos a float --
    precio_limpio=match.group(1).replace(",", "")
    precio_final=float(precio_limpio)

    return f"{precio_final:.2f}"
  except Exception as e:
    print(f"[bold red]error critico: {str(e)}[/]")
    driver.save_screenshot("error_precio_estricto.png")
    return None

#╔═══════════════════════════════════════════════════════════╗
#║ Funcion para acceder a todos los drops de informacion     ║
#║  ° Extraccion individual del elemento                     ║
#║  ° Pausas entre cada interaccion con los elementos        ║
#║  ° Iteracion e unteraccion por cada elemento              ║
#╚═══════════════════════════════════════════════════════════╝
def expand_all(driver, timeout=20, pause_between=0.5, retry_attempts=3):
  xpath_icono=(
    "//div[contains(@class, 'row-down-propertie')]"
    "//em[contains(@class, 'icon-angle-down-af')]"
    "/ancestor::div[contains(@class, 'row-down-propertie')]"
  )
  try:
    elementos=WebDriverWait(driver, timeout).until(EC.presence_of_all_elements_located((By.XPATH, xpath_icono)))
  except TimeoutException:
    print("[bold yellow]no se encontraron elementos para expandir[/]")
    return False
  
  total=len(elementos)
  print(f"[bold cyan]encontrados {total} elementos colapsados[/]")

  for index in range(total):
    for attempt in range(1, retry_attempts+1):
      try:
        # -- Re-localizar elementos por indice --
        elementos=driver.find_elements(By.XPATH, xpath_icono)
        if index>=len(elementos):
          print(f"[bold yellow]elemento {index+1} ya no existe")
          break
        boton=elementos[index]

        # -- Scroll y compensacion header --
        driver.execute_script(
          "const header=document.querySelector('header') || { offsetHeight: 0 };"
          "arguments[0].scrollIntoView({behavior: 'auto', block: 'center'});"
          "window.scrollBy(0, -header.offsetHeight);",
          boton
        )
        
        # --Esperar a ser clickeable y clicar
        WebDriverWait(driver, timeout).until(EC.element_to_be_clickable((By.XPATH, xpath_icono)))
        ActionChains(driver).move_to_element(boton).pause(pause_between).click().perform()

        time.sleep(pause_between)
        break
      except (StaleElementReferenceException, ElementClickInterceptedException, TimeoutException) as e:
        if attempt<retry_attempts:
          print(f"[bold yellow]reintentando elemento {index+1}, intento {attempt}/{retry_attempts}...[/]")
          time.sleep(pause_between)
          continue
        else:
          print("[bold red]no se pudo expandir el elemento {index+1} tras {retry_attempts} intentos: {e}[/]")
  
  print(f"[bold cyan]proceso completado: {total} intentos realizados[/]")
  return True

#╔═══════════════════════════════════════════════════════════╗
#║ Funcion para extraer informacion de daños materiales      ║
#║  ° Extraccion individual de la informacion                ║
#╚═══════════════════════════════════════════════════════════╝
def get_daños_materiales(driver):
  try:
    contenedor_principal=WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.ID, "card-1")))

    # -- Encontramos la seccion de "Coberturas"
    seccion_coberturas=WebDriverWait(contenedor_principal, 10).until(
      lambda d: d.find_element(
        By.XPATH,
        ".//p[@class='subtitle mb-0' and contains(text(), 'Coberturas')]/ancestor::div[contains(@class, 'cont-servicios')]"
      )
    )

    # -- Extraemos todos los elementos de la descripcion --
    coberturas=seccion_coberturas.find_elements(By.XPATH, ".//div[contains(@class, 'centrado')]/p[@class='description']")

    # -- Procesamos los resultados --
    lista_coberturas=[]
    for elemento in coberturas:
      texto=elemento.text.strip()
      if texto and texto not in lista_coberturas:
        lista_coberturas.append(texto)

    return lista_coberturas
  except Exception as e:
    print(f"[bold red]error extrayendo coberturas: {str(e)}[/]")
    driver.save_screenshot("error_coberturas.png")
    return []
  
def extraer_tiempo_pago_robo(driver):
  try:
    seccion_tiempo_pago=WebDriverWait(driver, 20).until(
      EC.presence_of_element_located((
        By.XPATH,
        "//p[@class='subtitle mb-0' and contains(text(), 'Tiempo de pago de indemnización')]"
        "/ancestor::div[contains(@class, 'cont-servicios')]"
      ))
    )
    
    tiempo_pago=seccion_tiempo_pago.find_element(
      By.XPATH,
      ".//p[@class='description']"
    ).text.strip()
    
    return tiempo_pago
  
  except Exception as e:
    
    return None
  

def extraer_extension_responsabilidad_civil(driver):
  try:
    WebDriverWait(driver, 15).until(
      EC.visibility_of_element_located((
        By.XPATH, 
        "//p[contains(@class, 'subtitle') and contains(text(), 'Extensión de Responsabilidad Civil')]"
      ))
    )
    
    columna_contenido=driver.find_element(
      By.XPATH,
      "//p[@class='subtitle mb-0' and contains(text(), 'Extensión de Responsabilidad Civil')]"
      "/ancestor::div[contains(@class, 'cont-servicios')]"
      "//div[contains(@class, 'col-5')]"
    )
    
    elementos=columna_contenido.find_elements(By.XPATH, ".//p[@class='description']")
    
    return list({elem.text.strip() for elem in elementos if elem.text.strip()})
      
  except TimeoutException:
    print("[yellow]Sección de Responsabilidad Civil no encontrada - puede que no esté presente[/]")
    return []
  except NoSuchElementException:
    print("[yellow]Elementos específicos no encontrados en la sección[/]")
    return []
  except Exception as e:
    print(f"[red]Error inesperado: {str(e)}[/]")
    driver.save_screenshot("error_inesperado_responsabilidad.png")
    return []
    
def extraer_gastos_medicos(driver):
  try:
    seccion_gastos_medicos=WebDriverWait(driver, 20).until(
      EC.presence_of_element_located((
        By.XPATH,
        "//div[contains(@class, 'cont-servicios') and .//div[contains(@class, 'd-flex justify-content-center') and .//p[text()='Coberturas']]]"
      ))
    )
    
    coberturas=seccion_gastos_medicos.find_elements(
      By.XPATH,
      ".//div[contains(@class, 'col-5')]//p[@class='description']"
    )
    
    lista_coberturas=[elemento.text.strip() for elemento in coberturas if elemento.text.strip()]
    return list(set(lista_coberturas))
  
  except Exception as e:
      print(f"[red]Error extrayendo gastos médicos: {str(e)}[/]")
      driver.save_screenshot("error_gastos_medicos.png")
      return []
    


def json_to_excel(json_data, output_file):
  cotizacion_id=list(json_data.keys())[0]
  datos_generales=json_data[cotizacion_id]
  
  wb=Workbook()
  
  if 'ws' in wb.sheetnames:
      ws_general=wb['ws']
      wb.remove(ws_general)
  ws_general=wb.create_sheet("Detalles Generales")
  
  general_data={
    "ID": [cotizacion_id],
    "Versión Autocompara": [datos_generales["Versión Autocompara "]],
    "Carrocerías": [datos_generales["Carrocerías"]],
    "Tipo": [datos_generales["Tipo"]],
    "Descripción": [datos_generales["Desc"]],
    "Año Modelo": [datos_generales["Año Mod"]],
    "CP": [datos_generales["CP"]],
    "Ciudad": [datos_generales["Ciudad"]],
    "Entidad": [datos_generales["Entidad"]],
    "Género": [datos_generales["Género"]],
    "Edad": [datos_generales["Edad"]],
    "Fecha Nacimiento": [datos_generales["Fecha Nacimiento"]],
  }
  
  aseguradoras=["CHUBB", "HDI", "ATLAS", "ZURICH", "MAPFRE", "GNP", "INBURSA", "ANA", "AIG", "QUÁLITAS", "AXA"]
  
  for aseguradora in aseguradoras:
    if aseguradora in datos_generales:
      precio=datos_generales[aseguradora]["precio"]
      general_data[aseguradora]=[float(precio)]
    else:
      general_data[aseguradora]=[None]
  
  df_general=pd.DataFrame(general_data)
  for r_idx, row in enumerate(dataframe_to_rows(df_general, index=False, header=True), 1):
    ws_general.append(row)
  
  for cell in ws_general[1]:
    cell.font=cell.font.copy(bold=True)
  
  ws_resumen=wb.create_sheet("Resumen Coberturas")
  
  resumen_data=[]
  for aseguradora in aseguradoras:
    if aseguradora in datos_generales:
      datos_aseguradora=datos_generales[aseguradora]
      resumen_data.append({
        "ID": cotizacion_id,
        "Aseguradora": aseguradora,
        "daños_materiales": len(datos_aseguradora["daños_materiales"]),
        "responsabilidad_civil": len(datos_aseguradora["responsabilidad_civil"]),
        "gastos_medicos": len(datos_aseguradora["gastos_medicos"]),
        "tiempo_robo": datos_aseguradora["tiempo_robo"] if datos_aseguradora["tiempo_robo"] else None
      })
  
  df_resumen=pd.DataFrame(resumen_data)
  for r_idx, row in enumerate(dataframe_to_rows(df_resumen, index=False, header=True), 1):
    ws_resumen.append(row)
  
  for cell in ws_resumen[1]:
    cell.font=cell.font.copy(bold=True)
  
  ws_danios=wb.create_sheet("Daños Materiales")
  
  danios_data=[]
  for aseguradora in aseguradoras:
    if aseguradora in datos_generales:
      danios=datos_generales[aseguradora]["daños_materiales"]
      for cobertura in danios:
        danios_data.append({
          "ID": cotizacion_id,
          "Aseguradora": aseguradora,
          "Cobertura": cobertura
        })
  
  df_danios=pd.DataFrame(danios_data)
  for r_idx, row in enumerate(dataframe_to_rows(df_danios, index=False, header=True), 1):
    ws_danios.append(row)
  
  for cell in ws_danios[1]:
    cell.font=cell.font.copy(bold=True)
  
  ws_robo=wb.create_sheet("Tiempo Robo")
  
  robo_data=[]
  for aseguradora in aseguradoras:
    if aseguradora in datos_generales:
      tiempo=datos_generales[aseguradora]["tiempo_robo"]
      robo_data.append({
        "ID": cotizacion_id,
        "Aseguradora": aseguradora,
        "Tiempo Robo": tiempo if tiempo else "No especificado"
      })
  
  df_robo=pd.DataFrame(robo_data)
  for r_idx, row in enumerate(dataframe_to_rows(df_robo, index=False, header=True), 1):
    ws_robo.append(row)
  
  for cell in ws_robo[1]:
    cell.font=cell.font.copy(bold=True)
  
  ws_responsabilidad=wb.create_sheet("Responsabilidad Civil")
  
  responsabilidad_data=[]
  for aseguradora in aseguradoras:
    if aseguradora in datos_generales:
      coberturas=datos_generales[aseguradora]["responsabilidad_civil"]
      for cobertura in coberturas:
        responsabilidad_data.append({
          "ID": cotizacion_id,
          "Aseguradora": aseguradora,
          "Cobertura": cobertura
        })
  
  df_responsabilidad=pd.DataFrame(responsabilidad_data)
  for r_idx, row in enumerate(dataframe_to_rows(df_responsabilidad, index=False, header=True), 1):
    ws_responsabilidad.append(row)
  
  for cell in ws_responsabilidad[1]:
    cell.font=cell.font.copy(bold=True)
  
  ws_gastos=wb.create_sheet("Gastos Médicos")
  
  gastos_data=[]
  for aseguradora in aseguradoras:
    if aseguradora in datos_generales:
      coberturas=datos_generales[aseguradora]["gastos_medicos"]
      for cobertura in coberturas:
        gastos_data.append({
          "ID": cotizacion_id,
          "Aseguradora": aseguradora,
          "Cobertura": cobertura
        })
  
  df_gastos=pd.DataFrame(gastos_data)
  for r_idx, row in enumerate(dataframe_to_rows(df_gastos, index=False, header=True), 1):
    ws_gastos.append(row)
  
  for cell in ws_gastos[1]:
    cell.font=cell.font.copy(bold=True)
  
  if "Sheet" in wb.sheetnames:
    wb.remove(wb["Sheet"])
  
  def autoajustar_columnas(ws):
    for column in ws.columns:
      max_length=0
      column_letter=column[0].column_letter
      for cell in column:
        try:
          if len(str(cell.value)) > max_length:
            max_length=len(str(cell.value))
        except:
          pass
      adjusted_width=(max_length + 2)
      ws.column_dimensions[column_letter].width=adjusted_width
  
  for sheet in wb.sheetnames:
    autoajustar_columnas(wb[sheet])
  
  wb.save(output_file)
  return output_file

#╔═══════════════════════════════════════════════════════════╗
#║ Funcion para extraer informacion de daños por auto        ║
#║  ° Extraccion individual de la informacion                ║
#╚═══════════════════════════════════════════════════════════╝
def get_aseguradoras(driver):
  drop_options(driver)
  aseguradoras=get_options(driver)
  return aseguradoras

def extract_informacion(driver, aseguradoras):
  id_data=str(uuid.uuid4())
  total_data={}
  for i, elemento in enumerate(aseguradoras):
    if i!=0:
      drop_options(driver)
      change_option(driver, elemento[0])
    precio=get_price(driver)
    expand_all(driver)
    data=get_daños_materiales(driver)
    total_data[elemento[1]]={
      "precio": precio,
      "daños_materiales": data,
      "tiempo_robo": extraer_tiempo_pago_robo(driver),
      "responsabilidad_civil": extraer_extension_responsabilidad_civil(driver),
      "gastos_medicos": extraer_gastos_medicos(driver)
    }
  return id_data, total_data

"""
Ejemplo de uso
id_data, data=extract_informacion(driver, aseguradoras)
"""